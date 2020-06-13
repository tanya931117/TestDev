#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/6/8 21:27
# @Author  : tanya
# @File    : my_excel.py
# @Software: PyCharm

from openpyxl import Workbook,load_workbook

class MyExcel():
    def __init__(self,file_path):
        self.file_path = file_path

    def write_excel(self,sheet_name,data,is_default):
        wb = Workbook()
        if is_default:
            ws = wb.active
            ws.title=sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        for index_row in range(len(data)):
            # ws.append(range(600))
            for index_col in range(len(data[index_row])):
                # print(type(data[index_row][index_col]))
                ws.cell(column=index_col+1,row=index_row+1,value=data[index_row][index_col])
        wb.save(self.file_path)

    def read_excel(self,sheet_name):
        wb = load_workbook(self.file_path)
        ws = wb[sheet_name]
        start_row = ws.min_row
        end_row = ws.max_row
        start_col = ws.min_column
        end_col = ws.max_column
        result = []
        for r in range(start_row,end_row+1):
            row = []
            for c in range(start_col,end_col+1):
                row.append(ws.cell(row=r, column=c).value)
            result.append(row)
        return result

if __name__=="__main__":
    me = MyExcel("demo1.xlsx")
    # me.write_excel("default",[[1,2,3],[4,5,6],[7,8,9]],True)
    res = me.read_excel("test")
    print(res)